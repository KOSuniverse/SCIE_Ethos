# PY Files/phase4_modeling/policy_export.py
# Phase 4C: Buy list / policy change XLSX export

from __future__ import annotations
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class PolicyExportManager:
    """
    Phase 4C: Generate Buy list / policy change XLSX exports.
    Creates comprehensive Excel reports for inventory policy recommendations.
    """
    
    def __init__(self):
        self.output_dir = Path("04_Data/Models/exports")
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def create_buy_list_export(
        self,
        policy_analyses: List[Dict[str, Any]],
        export_name: str = None
    ) -> str:
        """
        Create comprehensive buy list Excel export.
        
        Args:
            policy_analyses: List of comprehensive policy analysis results
            export_name: Optional custom export filename
        
        Returns:
            str: Path to generated Excel file
        """
        
        if not export_name:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_name = f"buy_list_policy_export_{timestamp}.xlsx"
        
        output_path = self.output_dir / export_name
        
        # Prepare data for different sheets
        buy_list_data = []
        policy_changes_data = []
        forecast_summary_data = []
        backtest_data = []
        
        for analysis in policy_analyses:
            part_number = analysis.get("part_number", "Unknown")
            current = analysis.get("current_analysis", {})
            rop = analysis.get("reorder_point", {})
            par = analysis.get("par_level", {})
            safety_stock = analysis.get("safety_stock", {})
            forecast = analysis.get("demand_forecast", {})
            backtest = analysis.get("backtest_performance", {})
            recommendations = analysis.get("recommendations", [])
            
            # Buy List Data
            current_stock = current.get("current_stock", 0)
            reorder_point = rop.get("reorder_point", 0)
            par_level = par.get("par_level", 0)
            
            # Determine action needed
            if current_stock <= reorder_point:
                action = "URGENT ORDER"
                qty_needed = par_level - current_stock
                priority = "High"
            elif current_stock < par_level * 0.8:
                action = "ORDER SOON"
                qty_needed = par_level - current_stock
                priority = "Medium"
            else:
                action = "MONITOR"
                qty_needed = 0
                priority = "Low"
            
            buy_list_data.append({
                "Part Number": part_number,
                "Current Stock": current_stock,
                "Reorder Point": reorder_point,
                "Par Level": par_level,
                "Safety Stock": safety_stock.get("safety_stock", 0),
                "Qty Needed": max(0, qty_needed),
                "Action": action,
                "Priority": priority,
                "Days of Supply": current.get("days_of_supply", 0),
                "Avg Demand": current.get("avg_demand", 0),
                "Lead Time": safety_stock.get("parameters", {}).get("lead_time_days", 0),
                "Service Level": analysis.get("service_level", 0.95),
                "Stock Status": current.get("stock_status", "Unknown")
            })
            
            # Policy Changes Data
            policy_changes_data.append({
                "Part Number": part_number,
                "Current ROP": "TBD",  # Would need current policy data
                "New ROP": reorder_point,
                "Current Par": "TBD",
                "New Par": par_level,
                "Current SS": "TBD",
                "New SS": safety_stock.get("safety_stock", 0),
                "Change Reason": "Optimized based on demand analysis",
                "Expected Impact": f"Improved service level to {analysis.get('service_level', 0.95)*100:.0f}%",
                "Implementation Date": datetime.now().strftime("%Y-%m-%d"),
                "Review Date": (datetime.now() + timedelta(days=90)).strftime("%Y-%m-%d")
            })
            
            # Forecast Summary Data
            if forecast:
                forecast_periods = len(forecast.get("forecast", []))
                avg_forecast = sum(forecast.get("forecast", [])) / max(1, forecast_periods)
                
                forecast_summary_data.append({
                    "Part Number": part_number,
                    "Forecast Method": forecast.get("method", "Unknown"),
                    "Forecast Periods": forecast_periods,
                    "Avg Forecasted Demand": avg_forecast,
                    "Current Avg Demand": current.get("avg_demand", 0),
                    "Demand Change %": ((avg_forecast - current.get("avg_demand", 0)) / max(current.get("avg_demand", 0), 1)) * 100,
                    "Forecast Confidence": "High" if backtest.get("accuracy_metrics", {}).get("mape", 100) < 20 else "Medium",
                    "Model Accuracy (MAPE)": backtest.get("accuracy_metrics", {}).get("mape", 0)
                })
            
            # Backtest Data
            if backtest and backtest.get("backtest_results"):
                bt_results = backtest["backtest_results"]
                for i, (actual, forecast_val, error) in enumerate(zip(
                    bt_results.get("actuals", []),
                    bt_results.get("forecasts", []),
                    bt_results.get("errors", [])
                )):
                    backtest_data.append({
                        "Part Number": part_number,
                        "Test Period": i + 1,
                        "Actual Demand": actual,
                        "Forecasted Demand": forecast_val,
                        "Absolute Error": error,
                        "Error %": (error / max(actual, 1)) * 100
                    })
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Buy List (Primary)
            if buy_list_data:
                buy_df = pd.DataFrame(buy_list_data)
                buy_df.to_excel(writer, sheet_name='Buy_List', index=False)
                self._format_buy_list_sheet(writer.book['Buy_List'], buy_df)
            
            # Sheet 2: Policy Changes
            if policy_changes_data:
                policy_df = pd.DataFrame(policy_changes_data)
                policy_df.to_excel(writer, sheet_name='Policy_Changes', index=False)
                self._format_policy_changes_sheet(writer.book['Policy_Changes'])
            
            # Sheet 3: Forecast Summary
            if forecast_summary_data:
                forecast_df = pd.DataFrame(forecast_summary_data)
                forecast_df.to_excel(writer, sheet_name='Forecast_Summary', index=False)
            
            # Sheet 4: Backtest Results
            if backtest_data:
                backtest_df = pd.DataFrame(backtest_data)
                backtest_df.to_excel(writer, sheet_name='Backtest_Results', index=False)
            
            # Sheet 5: Executive Summary
            self._create_executive_summary_sheet(writer, policy_analyses, buy_list_data)
        
        return str(output_path)
    
    def _format_buy_list_sheet(self, worksheet, buy_df):
        """Format the Buy List sheet with conditional formatting and styling."""
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        urgent_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        order_soon_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
        monitor_fill = PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Apply conditional formatting based on Action column
        action_col = None
        for idx, cell in enumerate(worksheet[1], 1):
            if cell.value == "Action":
                action_col = idx
                break
        
        if action_col:
            for row in range(2, len(buy_df) + 2):
                action_cell = worksheet.cell(row=row, column=action_col)
                if action_cell.value == "URGENT ORDER":
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=col).fill = urgent_fill
                elif action_cell.value == "ORDER SOON":
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=col).fill = order_soon_fill
                elif action_cell.value == "MONITOR":
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=col).fill = monitor_fill
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_policy_changes_sheet(self, worksheet):
        """Format the Policy Changes sheet."""
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_executive_summary_sheet(self, writer, policy_analyses, buy_list_data):
        """Create an executive summary sheet."""
        
        # Calculate summary metrics
        total_parts = len(policy_analyses)
        urgent_orders = len([item for item in buy_list_data if item["Action"] == "URGENT ORDER"])
        order_soon = len([item for item in buy_list_data if item["Action"] == "ORDER SOON"])
        monitor_only = len([item for item in buy_list_data if item["Action"] == "MONITOR"])
        
        total_qty_needed = sum([item["Qty Needed"] for item in buy_list_data])
        avg_service_level = sum([analysis.get("service_level", 0.95) for analysis in policy_analyses]) / max(total_parts, 1)
        
        # Create summary data
        summary_data = [
            ["Metric", "Value", "Description"],
            ["Total Parts Analyzed", total_parts, "Number of parts in this analysis"],
            ["Urgent Orders Required", urgent_orders, "Parts requiring immediate ordering"],
            ["Order Soon", order_soon, "Parts that should be ordered within review period"],
            ["Monitor Only", monitor_only, "Parts with adequate stock levels"],
            ["Total Quantity Needed", f"{total_qty_needed:,.0f}", "Total units to order across all urgent/soon parts"],
            ["Average Service Level", f"{avg_service_level:.1%}", "Target service level across all parts"],
            ["Analysis Date", datetime.now().strftime("%Y-%m-%d %H:%M"), "When this analysis was performed"]
        ]
        
        # Create DataFrame and add to Excel
        summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
        summary_df.to_excel(writer, sheet_name='Executive_Summary', index=False)
        
        # Format the summary sheet
        worksheet = writer.book['Executive_Summary']
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_single_part_export(
        self,
        policy_analysis: Dict[str, Any],
        export_name: str = None
    ) -> str:
        """
        Create detailed export for a single part analysis.
        
        Args:
            policy_analysis: Single comprehensive policy analysis result
            export_name: Optional custom export filename
        
        Returns:
            str: Path to generated Excel file
        """
        
        part_number = policy_analysis.get("part_number", "Unknown")
        
        if not export_name:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_name = f"policy_analysis_{part_number}_{timestamp}.xlsx"
        
        output_path = self.output_dir / export_name
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Summary
            self._create_part_summary_sheet(writer, policy_analysis)
            
            # Sheet 2: Forecast Details
            if policy_analysis.get("demand_forecast"):
                self._create_forecast_details_sheet(writer, policy_analysis)
            
            # Sheet 3: Policy Calculations
            self._create_policy_calculations_sheet(writer, policy_analysis)
            
            # Sheet 4: Backtest Results
            if policy_analysis.get("backtest_performance"):
                self._create_backtest_details_sheet(writer, policy_analysis)
        
        return str(output_path)
    
    def _create_part_summary_sheet(self, writer, policy_analysis):
        """Create summary sheet for single part analysis."""
        
        part_number = policy_analysis.get("part_number", "Unknown")
        current = policy_analysis.get("current_analysis", {})
        recommendations = policy_analysis.get("recommendations", [])
        
        # Create summary data
        summary_data = [
            ["Parameter", "Current Value", "Recommended Value", "Notes"],
            ["Part Number", part_number, part_number, ""],
            ["Current Stock", current.get("current_stock", 0), "", "Units on hand"],
            ["Days of Supply", f"{current.get('days_of_supply', 0):.1f}", "", "Current inventory coverage"],
            ["Stock Status", current.get("stock_status", "Unknown"), "", "Current inventory status"],
            ["Average Demand", f"{current.get('avg_demand', 0):.2f}", "", "Historical average"],
            ["Demand Std Dev", f"{current.get('demand_std', 0):.2f}", "", "Demand variability"],
            ["Reorder Point", "", policy_analysis.get("reorder_point", {}).get("reorder_point", 0), "When to reorder"],
            ["Par Level", "", policy_analysis.get("par_level", {}).get("par_level", 0), "Maximum stock target"],
            ["Safety Stock", "", policy_analysis.get("safety_stock", {}).get("safety_stock", 0), "Buffer inventory"],
            ["Service Level", "", f"{policy_analysis.get('service_level', 0.95):.1%}", "Target service level"]
        ]
        
        # Add recommendations
        for i, rec in enumerate(recommendations[:5]):  # Limit to top 5
            summary_data.append([f"Recommendation {i+1}", rec, "", ""])
        
        # Create DataFrame and add to Excel
        summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_forecast_details_sheet(self, writer, policy_analysis):
        """Create detailed forecast sheet."""
        
        forecast = policy_analysis.get("demand_forecast", {})
        forecast_values = forecast.get("forecast", [])
        confidence_intervals = forecast.get("confidence_intervals", [])
        
        if not forecast_values:
            return
        
        # Create forecast data
        forecast_data = []
        for i, (forecast_val, ci) in enumerate(zip(forecast_values, confidence_intervals)):
            forecast_data.append({
                "Period": i + 1,
                "Forecast": forecast_val,
                "Lower Bound": ci[0] if ci else forecast_val * 0.8,
                "Upper Bound": ci[1] if ci else forecast_val * 1.2,
                "Confidence Width": (ci[1] - ci[0]) if ci else forecast_val * 0.4
            })
        
        forecast_df = pd.DataFrame(forecast_data)
        forecast_df.to_excel(writer, sheet_name='Forecast_Details', index=False)
    
    def _create_policy_calculations_sheet(self, writer, policy_analysis):
        """Create policy calculations details sheet."""
        
        # Extract calculation details
        safety_stock = policy_analysis.get("safety_stock", {})
        reorder_point = policy_analysis.get("reorder_point", {})
        par_level = policy_analysis.get("par_level", {})
        
        calc_data = [
            ["Calculation", "Formula", "Parameters", "Result"],
            ["Safety Stock", safety_stock.get("formula", ""), str(safety_stock.get("parameters", {})), safety_stock.get("safety_stock", 0)],
            ["Reorder Point", reorder_point.get("formula", ""), str(reorder_point.get("parameters", {})), reorder_point.get("reorder_point", 0)],
            ["Par Level", par_level.get("formula", ""), str(par_level.get("parameters", {})), par_level.get("par_level", 0)]
        ]
        
        calc_df = pd.DataFrame(calc_data[1:], columns=calc_data[0])
        calc_df.to_excel(writer, sheet_name='Policy_Calculations', index=False)
    
    def _create_backtest_details_sheet(self, writer, policy_analysis):
        """Create backtest results details sheet."""
        
        backtest = policy_analysis.get("backtest_performance", {})
        bt_results = backtest.get("backtest_results", {})
        
        if not bt_results:
            return
        
        # Create backtest data
        backtest_data = []
        for i, (actual, forecast_val, error) in enumerate(zip(
            bt_results.get("actuals", []),
            bt_results.get("forecasts", []),
            bt_results.get("errors", [])
        )):
            backtest_data.append({
                "Test Period": i + 1,
                "Actual": actual,
                "Forecast": forecast_val,
                "Error": error,
                "Error %": (error / max(actual, 1)) * 100,
                "Squared Error": error ** 2
            })
        
        backtest_df = pd.DataFrame(backtest_data)
        backtest_df.to_excel(writer, sheet_name='Backtest_Details', index=False)
